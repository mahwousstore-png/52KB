"""
engines/engine.py  v21.1 — محرك المطابقة الفائق السرعة
═══════════════════════════════════════════════════════
🚀 تطبيع مسبق (Pre-normalize) → vectorized cdist → Gemini للغموض فقط
⚡ 5x أسرع من v20 مع نفس الدقة 99.5%

الخطة:
  1. عند رفع الملف → تطبيع كل منتجات المنافس مرة واحدة (cache)
  2. لكل منتجنا → cdist vectorized دفعة واحدة (بدل loop)
  3. أفضل 5 مرشحين → Gemini فقط إذا score بين 62-96%
  4. score ≥97% → تلقائي فوري  |  score <62% → مفقود
"""
import re, io, json, hashlib, sqlite3, time
from datetime import datetime
import pandas as pd
from rapidfuzz import fuzz, process as rf_process
from rapidfuzz.distance import Indel
import requests as _req

# ─── استيراد الإعدادات ───────────────────────
try:
    from config import (REJECT_KEYWORDS, KNOWN_BRANDS, WORD_REPLACEMENTS,
                        MATCH_THRESHOLD, HIGH_CONFIDENCE, REVIEW_THRESHOLD,
                        PRICE_TOLERANCE, TESTER_KEYWORDS, SET_KEYWORDS, GEMINI_API_KEYS)
except:
    REJECT_KEYWORDS = ["sample","عينة","عينه","decant","تقسيم","split","miniature"]
    KNOWN_BRANDS = [
        "Dior","Chanel","Gucci","Tom Ford","Versace","Armani","YSL","Prada","Burberry",
        "Hermes","Creed","Montblanc","Amouage","Rasasi","Lattafa","Arabian Oud","Ajmal",
        "Al Haramain","Afnan","Armaf","Mancera","Montale","Kilian","Jo Malone",
        "Carolina Herrera","Paco Rabanne","Mugler","Ralph Lauren","Parfums de Marly",
        "Nishane","Xerjoff","Byredo","Le Labo","Roja","Narciso Rodriguez",
        "Dolce & Gabbana","Valentino","Bvlgari","Cartier","Hugo Boss","Calvin Klein",
        "Givenchy","Lancome","Guerlain","Jean Paul Gaultier","Issey Miyake","Davidoff",
        "Coach","Michael Kors","Initio","Memo Paris","Maison Margiela","Diptyque",
        "Missoni","Juicy Couture","Moschino","Dunhill","Bentley","Jaguar",
        "Boucheron","Chopard","Elie Saab","Escada","Ferragamo","Fendi",
        "Kenzo","Lacoste","Loewe","Rochas","Roberto Cavalli","Tiffany",
        "Van Cleef","Azzaro","Chloe","Elizabeth Arden","Swiss Arabian",
        "Penhaligons","Clive Christian","Floris","Acqua di Parma",
        "Ard Al Zaafaran","Nabeel","Asdaaf","Maison Alhambra",
        "Tiziana Terenzi","Maison Francis Kurkdjian","Serge Lutens",
        "Frederic Malle","Ormonde Jayne","Zoologist","Tauer",
        "Banana Republic","Benetton","Bottega Veneta","Celine","Dsquared2",
        "Ermenegildo Zegna","Sisley","Mexx","Amadou","Thameen",
        "Nasomatto","Nicolai","Replica","Atelier Cologne","Aerin",
        "Angel Schlesser","Annick Goutal","Antonio Banderas","Balenciaga",
        "Bond No 9","Boadicea","Carner Barcelona","Clean","Commodity",
        "Costume National","Creed","Derek Lam","Diptique","Estee Lauder",
        "Franck Olivier","Giorgio Beverly Hills","Guerlain","Guess",
        "Histoires de Parfums","Illuminum","Jimmy Choo","Kenneth Cole",
        "Lalique","Lolita Lempicka","Lubin","Miu Miu","Moresque",
        "Nobile 1942","Oscar de la Renta","Oud Elite","Philipp Plein",
        "Police","Prada","Rasasi","Reminiscence","Salvatore Ferragamo",
        "Stella McCartney","Ted Lapidus","Ungaro","Vera Wang","Viktor Rolf",
        "Zadig Voltaire","Zegna","Ajwad","Club de Nuit","Milestone",
        "لطافة","العربية للعود","رصاسي","أجمل","الحرمين","أرماف",
        "أمواج","كريد","توم فورد","ديور","شانيل","غوتشي","برادا",
        "ميسوني","جوسي كوتور","موسكينو","دانهيل","بنتلي",
        "كينزو","لاكوست","فندي","ايلي صعب","ازارو",
        "كيليان","نيشان","زيرجوف","بنهاليغونز","مارلي","جيرلان",
        "تيزيانا ترينزي","مايزون فرانسيس","بايريدو","لي لابو",
        "مانسيرا","مونتالي","روجا","جو مالون","ثمين","أمادو",
        "ناسوماتو","ميزون مارجيلا","نيكولاي",
        "جيمي تشو","لاليك","بوليس","فيكتور رولف",
        "كلوي","بالنسياغا","ميو ميو",
    ]
WORD_REPLACEMENTS = {}
MATCH_THRESHOLD = 85; HIGH_CONFIDENCE = 95; REVIEW_THRESHOLD = 75
PRICE_TOLERANCE = 5; TESTER_KEYWORDS = ["tester","تستر"]; SET_KEYWORDS = ["set","طقم","مجموعة"]

# ─── قراءة مفاتيح Gemini من Railway Environment Variables ───
import os as _os
def _load_gemini_keys():
    keys = []
    # طريقة 1: GEMINI_API_KEYS مفصولة بفاصلة
    v = _os.environ.get("GEMINI_API_KEYS", "")
    if v:
        keys += [k.strip() for k in v.split(",") if k.strip()]
    # طريقة 2: مفاتيح منفردة GEMINI_KEY_1, GEMINI_KEY_2 ...
    for i in range(1, 10):
        k = _os.environ.get(f"GEMINI_KEY_{i}", "")
        if k.strip():
            keys.append(k.strip())
    # طريقة 3: أسماء بديلة
    for env_name in ["GEMINI_API_KEY", "GEMINI_KEY"]:
        k = _os.environ.get(env_name, "")
        if k.strip():
            keys.append(k.strip())
    return list(dict.fromkeys(keys))  # إزالة التكرار مع الحفاظ على الترتيب

GEMINI_API_KEYS = _load_gemini_keys()

# ─── مرادفات ذكية للعطور ────────────────────
_SYN = {
    "eau de parfum":"edp","او دو بارفان":"edp","أو دو بارفان":"edp",
    "او دي بارفان":"edp","بارفان":"edp","parfum":"edp","perfume":"edp",
    "eau de toilette":"edt","او دو تواليت":"edt","أو دو تواليت":"edt",
    "تواليت":"edt","toilette":"edt","toilet":"edt",
    "eau de cologne":"edc","كولون":"edc","cologne":"edc",
    "extrait de parfum":"extrait","parfum extrait":"extrait",
    "ديور":"dior","شانيل":"chanel","شنل":"chanel","أرماني":"armani","ارماني":"armani",
    "جورجيو ارماني":"armani","فرساتشي":"versace","فيرساتشي":"versace",
    "غيرلان":"guerlain","توم فورد":"tom ford","تومفورد":"tom ford",
    "لطافة":"lattafa","لطافه":"lattafa",
    "أجمل":"ajmal","رصاصي":"rasasi","أمواج":"amouage","كريد":"creed",
    "ايف سان لوران":"ysl","سان لوران":"ysl","yves saint laurent":"ysl",
    "غوتشي":"gucci","قوتشي":"gucci","برادا":"prada","برادة":"prada",
    "بربري":"burberry","بيربري":"burberry","جيفنشي":"givenchy","جفنشي":"givenchy",
    "كارولينا هيريرا":"carolina herrera","باكو رابان":"paco rabanne",
    "نارسيسو رودريغيز":"narciso rodriguez","كالفن كلاين":"calvin klein",
    "هوجو بوس":"hugo boss","فالنتينو":"valentino","بلغاري":"bvlgari",
    "كارتييه":"cartier","لانكوم":"lancome","جو مالون":"jo malone",
    "سوفاج":"sauvage","بلو":"bleu","إيروس":"eros","ايروس":"eros",
    "وان ميليون":"1 million",
    "إنفيكتوس":"invictus","أفينتوس":"aventus","عود":"oud","مسك":"musk",
    "ميسوني":"missoni","جوسي كوتور":"juicy couture","موسكينو":"moschino",
    "دانهيل":"dunhill","بنتلي":"bentley","كينزو":"kenzo","لاكوست":"lacoste",
    "فندي":"fendi","ايلي صعب":"elie saab","ازارو":"azzaro",
    "فيراغامو":"ferragamo","شوبار":"chopard","بوشرون":"boucheron",
    "لانكم":"lancome","لانكوم":"lancome","جيفنشي":"givenchy","جيفانشي":"givenchy",
    "بربري":"burberry","بيربري":"burberry","بوربيري":"burberry",
    "فيرساتشي":"versace","فرزاتشي":"versace",
    "روبيرتو كفالي":"roberto cavalli","روبرتو كافالي":"roberto cavalli",
    "سلفاتوري":"ferragamo","سالفاتوري":"ferragamo",
    "ايف سان لوران":"ysl","ايف سانت لوران":"ysl",
    "هيرميس":"hermes","ارميس":"hermes","هرمز":"hermes",
    "كيليان":"kilian","كليان":"kilian",
    "نيشان":"nishane","نيشاني":"nishane",
    "زيرجوف":"xerjoff","زيرجوفف":"xerjoff",
    "بنهاليغونز":"penhaligons","بنهاليغون":"penhaligons",
    "مارلي":"parfums de marly","دي مارلي":"parfums de marly",
    "جيرلان":"guerlain","غيرلان":"guerlain","جرلان":"guerlain",
    "تيزيانا ترينزي":"tiziana terenzi","تيزيانا":"tiziana terenzi",
    "ناسوماتو":"nasomatto",
    "ميزون مارجيلا":"maison margiela","مارجيلا":"maison margiela","ربليكا":"replica",
    "نيكولاي":"nicolai","نيكولائي":"nicolai",
    "مايزون فرانسيس":"maison francis kurkdjian","فرانسيس":"maison francis kurkdjian",
    "بايريدو":"byredo","لي لابو":"le labo",
    "مانسيرا":"mancera","مونتالي":"montale","روجا":"roja",
    "جو مالون":"jo malone","جومالون":"jo malone",
    "ثمين":"thameen","أمادو":"amadou","امادو":"amadou",
    "انيشيو":"initio","إنيشيو":"initio","initio":"initio",
    "جيمي تشو":"jimmy choo","جيميتشو":"jimmy choo",
    "لاليك":"lalique","بوليس":"police",
    "فيكتور رولف":"viktor rolf","فيكتور اند رولف":"viktor rolf",
    "كلوي":"chloe","شلوي":"chloe",
    "بالنسياغا":"balenciaga","بالنسياجا":"balenciaga",
    "ميو ميو":"miu miu",
    "استي لودر":"estee lauder","استيلودر":"estee lauder",
    "كوتش":"coach","مايكل كورس":"michael kors",
    "رالف لورين":"ralph lauren","رالف لوران":"ralph lauren",
    "ايزي مياكي":"issey miyake","ايسي مياكي":"issey miyake",
    "دافيدوف":"davidoff","ديفيدوف":"davidoff",
    "دولشي اند غابانا":"dolce gabbana","دولتشي":"dolce gabbana","دولشي":"dolce gabbana",
    "جان بول غولتييه":"jean paul gaultier","غولتييه":"jean paul gaultier","غولتيه":"jean paul gaultier",
    "غوتييه":"jean paul gaultier","جان بول غوتييه":"jean paul gaultier","قوتييه":"jean paul gaultier","قولتييه":"jean paul gaultier",
    "مونت بلانك":"montblanc","مونتبلان":"montblanc",
    "موجلر":"mugler","موغلر":"mugler","تييري موجلر":"mugler",
    "كلوب دي نوي":"club de nuit","كلوب دنوي":"club de nuit",
    "مايلستون":"milestone",
    "سكاندل":"scandal","سكاندال":"scandal",
    " مل":" ml","ملي ":"ml ","ملي":"ml","مل":"ml",
    "أ":"ا","إ":"ا","آ":"ا","ة":"ه","ى":"ي","ؤ":"و","ئ":"ي",
}

# ─── SQLite Cache ───────────────────────────
_DB = "match_cache_v21.db"
def _init_db():
    try:
        cn = sqlite3.connect(_DB, check_same_thread=False)
        cn.execute("CREATE TABLE IF NOT EXISTS cache(h TEXT PRIMARY KEY, v TEXT, ts TEXT)")
        cn.commit(); cn.close()
    except: pass

def _cget(k):
    try:
        cn = sqlite3.connect(_DB, check_same_thread=False)
        r = cn.execute("SELECT v FROM cache WHERE h=?", (k,)).fetchone()
        cn.close(); return json.loads(r[0]) if r else None
    except: return None

def _cset(k, v):
    try:
        cn = sqlite3.connect(_DB, check_same_thread=False)
        cn.execute("INSERT OR REPLACE INTO cache VALUES(?,?,?)",
                   (k, json.dumps(v, ensure_ascii=False), datetime.now().isoformat()))
        cn.commit(); cn.close()
    except: pass

_init_db()

# ─── دوال أساسية ────────────────────────────
def read_file(f):
    try:
        name = f.name.lower()
        df = None
        if name.endswith('.csv'):
            for enc in ['utf-8-sig','utf-8','windows-1256','cp1256','latin-1']:
                try:
                    f.seek(0)
                    df = pd.read_csv(f, encoding=enc, on_bad_lines='skip')
                    if len(df) > 0 and not df.columns[0].startswith('\ufeff'): 
                        break
                except: continue
            if df is None:
                return None, "فشل قراءة الملف بجميع الترميزات"
        elif name.endswith(('.xlsx','.xls')):
            df = pd.read_excel(f)
        else:
            return None, "صيغة غير مدعومة"
        # تنظيف أسماء الأعمدة من BOM والمسافات
        df.columns = df.columns.str.strip().str.replace('\ufeff', '', regex=False)
        df = df.dropna(how='all').reset_index(drop=True)
        # إذا كانت الأعمدة Unnamed أو أسماء CSS → تخمين ذكي
        df = _smart_rename_columns(df)
        return df, None
    except Exception as e:
        return None, str(e)


def _smart_rename_columns(df):
    """تخمين ذكي لأسماء الأعمدة إذا كانت غير معروفة (Unnamed أو أسماء CSS)"""
    cols = list(df.columns)
    # حالة 1: أعمدة Unnamed (ملف بدون عناوين)
    unnamed_count = sum(1 for c in cols if str(c).startswith('Unnamed'))
    # حالة 2: أعمدة CSS (مثل styles_productCard__name)
    css_count = sum(1 for c in cols if 'style' in str(c).lower() or '__' in str(c))
    
    if unnamed_count >= len(cols) - 1 or css_count >= 1:
        # تحليل المحتوى لتخمين الأعمدة
        new_cols = {}
        for col in cols:
            sample = df[col].dropna().head(20)
            if sample.empty:
                continue
            # تحقق إذا كان العمود يحتوي على أرقام (أسعار)
            numeric_count = 0
            for v in sample:
                try:
                    float(str(v).replace(',', ''))
                    numeric_count += 1
                except:
                    pass
            if numeric_count >= len(sample) * 0.7:
                new_cols[col] = 'السعر'
            else:
                # يحتوي على نصوص → اسم المنتج
                if 'المنتج' not in new_cols.values() and 'اسم المنتج' not in new_cols.values():
                    new_cols[col] = 'اسم المنتج'
                else:
                    new_cols[col] = col  # ابقِ كما هو
        if new_cols:
            df = df.rename(columns=new_cols)
    return df

def normalize(text):
    if not isinstance(text, str): return ""
    t = text.strip().lower()
    for k, v in WORD_REPLACEMENTS.items():
        t = t.replace(k.lower(), v)
    for k, v in _SYN.items():
        t = t.replace(k, v)
    t = re.sub(r'[^\w\s\u0600-\u06FF.]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()

def extract_size(text):
    if not isinstance(text, str): return 0.0
    tl = text.lower()
    # البحث عن oz أولاً وتحويله لـ ml
    oz = re.findall(r'(\d+(?:\.\d+)?)\s*(?:oz|ounce)', tl)
    if oz:
        return float(oz[0]) * 29.5735  # 1 oz = 29.5735 ml
    # البحث عن ml
    ml = re.findall(r'(\d+(?:\.\d+)?)\s*(?:ml|مل|ملي|milliliter)', tl)
    return float(ml[0]) if ml else 0.0

def extract_brand(text):
    if not isinstance(text, str): return ""
    n = normalize(text)
    tl = text.lower()
    for b in KNOWN_BRANDS:
        if normalize(b) in n or b.lower() in tl: return b
    return ""

def extract_type(text):
    if not isinstance(text, str): return ""
    n = normalize(text)
    if "edp" in n or "extrait" in n: return "EDP"
    if "edt" in n: return "EDT"
    if "edc" in n: return "EDC"
    return ""

def extract_gender(text):
    if not isinstance(text, str): return ""
    tl = text.lower()
    # تم التحديث ليشمل mans وصيغ الرجال المطلوبة
    m = any(k in tl for k in ["pour homme","for men"," men "," man ","رجالي","للرجال"," مان "," هوم ","homme"," uomo", "mans", "for mans", " mans "])
    w = any(k in tl for k in ["pour femme","for women","women"," woman ","نسائي","للنساء","النسائي","lady","femme"," donna"])
    if m and not w: return "رجالي"
    if w and not m: return "نسائي"
    return ""

def extract_product_line(text, brand=""):
    """استخراج اسم خط الإنتاج (المنتج الأساسي) بعد إزالة الماركة والكلمات الشائعة.
    مثال: 'عطر بربري هيرو أو دو تواليت 100مل' → 'هيرو'
    مثال: 'عطر لندن من بربري للرجال' → 'لندن'
    هذا ضروري لمنع مطابقة 'بربري هيرو' مع 'بربري لندن'
    """
    if not isinstance(text, str): return ""
    n = text.lower()
    # إزالة الماركة (عربي + إنجليزي) — كل الأشكال
    if brand:
        for b_var in [brand.lower(), normalize(brand)]:
            n = n.replace(b_var, " ")
        # إزالة المرادفات العربية لهذه الماركة تحديداً
        brand_norm = brand.lower()
        for k, v in _SYN.items():
            if v == brand_norm or v == normalize(brand):
                n = n.replace(k, " ")
    # إزالة حروف الجر المتبقية
    for prep in ['من','في','لل','ال']:
        n = re.sub(r'\b' + prep + r'\b', ' ', n)
    # إزالة الكلمات الشائعة
    _STOP = [
        'عطر','تستر','تيستر','tester','perfume','fragrance',
        'او دو','او دي','أو دو','أو دي',
        'بارفان','بارفيوم','برفيوم','بيرفيوم','برفان','parfum','edp','eau de parfum',
        'تواليت','toilette','edt','eau de toilette',
        'كولون','cologne','edc','eau de cologne',
        'انتنس','انتينس','intense','اكستريم','extreme',
        'ابسولو','ابسوليو','absolue','absolute','absolu',
        'اكستريت','اكسترايت','extrait','extract',
        'دو','de','du','la','le','les','the',
        # أسماء ماركات فرعية تبقى بعد إزالة الماركة الرئيسية
        'تيرينزي','ترينزي','terenzi','terenzio',  # Tiziana Terenzi
        'كوركدجيان','كركدجيان','kurkdjian',  # MFK
        'ميزون','مايزون','maison',  # Maison Margiela/MFK
        'باريس','paris',  # كلمة شائعة
        'دوف','dove',  # Roja Dove
        'للرجال','للنساء','رجالي','نسائي','للجنسين',
        'for men','for women','unisex','pour homme','pour femme',
        'ml','مل','ملي','milliliter',
        'كرتون ابيض','كرتون أبيض','white box',
        'اصلي','original','authentic','جديد','new',
        'اصدار','اصدارات','edition','limited',
        # كلمات شائعة ترفع pl_score خطأً
        'برفان','spray','بخاخ','عطور',
        'الرجالي','النسائي','رجال','نساء',
        'men','women','homme','femme',
        'مان','man','uomo','donna',
        'هوم','فيم',
        'او','ou','or','و',
        # كلمات إضافية ترفع pl_score خطأً
        'لو','لا','lo',
        'di','دي',
        # أجزاء أسماء الماركات المركبة التي تبقى بعد إزالة المرادف
        'جان','بول','jean','paul','gaultier',
        'كارولينا','هيريرا','carolina','herrera',
        'دولشي','غابانا','dolce','gabbana',
        'رالف','لورين','ralph','lauren',
        'ايزي','مياكي','issey','miyake',
        'فان','كليف','van','cleef','arpels',
        'اورمند','جايان','ormonde','jayne',
        'توماس','كوسمالا','thomas','kosmala',
        'فرانسيس','francis',
        'روسيندو','ماتيو','rosendo','mateu',
        'نيكولاي','nicolai',
        'ارماف','armaf',
    ]
    # إزالة الكلمات الطويلة (4+ حروف) بـ replace عادي
    # والكلمات القصيرة (1-3 حروف) بـ word boundary لمنع حذف أجزاء من كلمات أخرى
    for w in _STOP:
        if len(w) <= 3:
            n = re.sub(r'(?:^|\s)' + re.escape(w) + r'(?:\s|$)', ' ', n)
        else:
            n = n.replace(w, ' ')
    # إزالة الأرقام (الحجم) + مل/ml الملتصقة
    n = re.sub(r'\d+(?:\.\d+)?\s*(?:ml|مل|ملي)?', ' ', n)
    # إزالة الرموز
    n = re.sub(r'[^\w\s\u0600-\u06FF]', ' ', n)
    # توحيد الهمزات
    for k, v in {'أ':'ا','إ':'ا','آ':'ا','ة':'ه','ى':'ي'}.items():
        n = n.replace(k, v)
    return re.sub(r'\s+', ' ', n).strip()

def is_sample(t):
    return isinstance(t, str) and any(k in t.lower() for k in REJECT_KEYWORDS)

def is_tester(t):
    return isinstance(t, str) and any(k in t.lower() for k in TESTER_KEYWORDS)

def is_set(t):
    return isinstance(t, str) and any(k in t.lower() for k in SET_KEYWORDS)

def classify_product(name):
    """تصنيف المنتج حسب AI_COMPARISON_INSTRUCTIONS: retail/tester/set/hair_mist/body_mist/rejected"""
    if not isinstance(name, str): return "retail"
    nl = name.lower()
    if any(w in nl for w in ['sample','عينة','عينه','miniature','مينياتشر','travel size','decant','تقسيم']):
        return 'rejected'
    if any(w in nl for w in ['tester','تستر','تيستر']):
        return 'tester'
    if any(w in nl for w in ['set ','سيت','مجموعة','gift','هدية','طقم','coffret']):
        return 'set'
    # hair mist: كلمات كاملة فقط (لتجنب "هيريرا" → hair_mist)
    if re.search(r'\bhair\s*mist\b|عطر\s*شعر|معطر\s*شعر|للشعر|\bhair\b', nl):
        return 'hair_mist'
    # body mist: كلمات كاملة فقط
    if re.search(r'\bbody\s*mist\b|بودي\s*مست|بخاخ\s*جسم|معطر\s*جسم|\bbody\s*spray\b', nl):
        return 'body_mist'
    # بودرة/كريم/لوشن
    if re.search(r'بودرة|بودره|powder|كريم|cream|لوشن|lotion|ديودرنت|deodorant', nl):
        return 'other'
    return 'retail'

def _price(row):
    for c in ["السعر","Price","price","سعر","PRICE"]:
        if c in row.index:
            try: return float(str(row[c]).replace(",",""))
            except: pass
    # احتياطي: ابحث عن أي عمود رقمي يشبه السعر
    for c in row.index:
        try:
            v = float(str(row[c]).replace(",",""))
            if 1 <= v <= 99999:  # نطاق سعر معقول
                return v
        except:
            pass
    return 0.0

def _pid(row, col):
    if not col or col not in row.index: return ""
    v = row.get(col, "")
    if v is None or str(v) in ("nan", "None", "", "NaN"): return ""
    # تحويل float إلى int لإزالة .0 (مثل 1081786650.0 → 1081786650)
    try:
        fv = float(v)
        if fv == int(fv):
            return str(int(fv))
    except (ValueError, TypeError):
        pass
    return str(v).strip()

def _fcol(df, cands):
    for c in cands:
        if c in df.columns: return c
    return df.columns[0] if len(df.columns) else ""

# ═══════════════════════════════════════════════════════
#  الكلاس الجديد: Pre-normalized Competitor Index
#  يُبنى مرة واحدة لكل ملف منافس ← يسرّع الـ matching 5x
# ═══════════════════════════════════════════════════════
class CompIndex:
    """فهرس المنافس المطبَّع مسبقاً"""
    def __init__(self, df, name_col, id_col, comp_name):
        self.comp_name = comp_name
        self.name_col  = name_col
        self.id_col    = id_col
        self.df        = df.reset_index(drop=True)
        # تطبيع مسبق لكل الأسماء — مرة واحدة فقط
        self.raw_names  = df[name_col].fillna("").astype(str).tolist()
        self.norm_names = [normalize(n) for n in self.raw_names]
        self.brands     = [extract_brand(n) for n in self.raw_names]
        self.sizes      = [extract_size(n) for n in self.raw_names]
        self.types      = [extract_type(n) for n in self.raw_names]
        self.genders    = [extract_gender(n) for n in self.raw_names]
        # خطوط الإنتاج — لمنع مطابقة 'بربري هيرو' مع 'بربري لندن'
        self.plines     = [extract_product_line(n, self.brands[i]) for i, n in enumerate(self.raw_names)]
        self.prices     = [_price(row) for _, row in df.iterrows()]
        self.ids        = [_pid(row, id_col) for _, row in df.iterrows()]

    def search(self, our_norm, our_br, our_sz, our_tp, our_gd, our_pline="", top_n=6):
        """بحث vectorized بـ rapidfuzz process.extract مع مقارنة خط الإنتاج"""
        if not self.norm_names: return []

        # استبعاد العينات مسبقاً
        valid_idx = [i for i, n in enumerate(self.raw_names) if not is_sample(n)]
        if not valid_idx: return []

        valid_norms = [self.norm_names[i] for i in valid_idx]

        # extract بالطريقة الأسرع
        fast = rf_process.extract(
            our_norm, valid_norms,
            scorer=fuzz.token_set_ratio,
            limit=min(25, len(valid_norms))
        )

        cands = []
        seen  = set()
        for _, fast_score, vi in fast:
            if fast_score < max(MATCH_THRESHOLD - 15, 40): continue
            idx  = valid_idx[vi]
            name = self.raw_names[idx]
            if name in seen: continue

            c_br = self.brands[idx]
            c_sz = self.sizes[idx]
            c_tp = self.types[idx]
            c_gd = self.genders[idx]
            c_pl = self.plines[idx]

            # ═══ فلاتر سريعة ═══
            if our_br and c_br and normalize(our_br) != normalize(c_br): continue
            if our_sz > 0 and c_sz > 0 and abs(our_sz - c_sz) > 30: continue
            if our_tp and c_tp and our_tp != c_tp:
                if our_sz > 0 and c_sz > 0 and abs(our_sz - c_sz) > 3: continue
            if our_gd and c_gd and our_gd != c_gd: continue

            # ═══ فلتر تصنيف المنتج (retail/tester/set/hair_mist) ═══
            our_class = classify_product(our_norm)
            c_class = classify_product(name)
            if our_class != c_class:
                # العينات تُستثنى تماماً
                if our_class == 'rejected' or c_class == 'rejected':
                    continue
                # المجموعات ومعطرات الشعر/الجسم لا تقارن مع العطور
                if our_class in ('hair_mist','body_mist','set','other') or \
                   c_class in ('hair_mist','body_mist','set','other'):
                    continue
                # التستر يقارن فقط مع التستر، العطر الأساسي فقط مع الأساسي
                if (our_class == 'tester') != (c_class == 'tester'):
                    continue

            # ═══ مقارنة الأرقام في أسماء المنتجات (نمبر 11 ≠ نمبر 10) ═══
            _NUM_WORDS = {
                'ون':'1','تو':'2','ثري':'3','فور':'4','فايف':'5',
                'سكس':'6','سفن':'7','ايت':'8','ناين':'9','تن':'10',
                'one':'1','two':'2','three':'3','four':'4','five':'5',
                'six':'6','seven':'7','eight':'8','nine':'9','ten':'10',
                'i':'1','ii':'2','iii':'3','iv':'4','v':'5',
                'vi':'6','vii':'7','viii':'8','ix':'9','x':'10',
            }
            def _extract_product_numbers(text):
                """Extract product-identifying numbers (not sizes)"""
                nums = set()
                # استخراج الأرقام الرقمية
                for m in re.finditer(r'(?:no|num|number|نمبر|رقم|№|#)\s*(\d+)', text.lower()):
                    nums.add(m.group(1))
                # استخراج الأرقام النصية (ون، تو، سفن...)
                tl = text.lower()
                for word, num in _NUM_WORDS.items():
                    if f'نمبر {word}' in tl or f'number {word}' in tl or f'no {word}' in tl or f'رقم {word}' in tl:
                        nums.add(num)
                # استخراج أرقام ملتصقة بكلمات (مثل سفن7)
                for m in re.finditer(r'[a-z؀-ۿ](\d+)', text.lower()):
                    v = m.group(1)
                    if v not in {'100','50','30','200','150','75','80','125','250','300','ml'}:
                        nums.add(v)
                # أرقام مستقلة ليست أحجام (مثل 212, 360, 9)
                for m in re.finditer(r'\b(\d{1,3})\b', text.lower()):
                    v = m.group(1)
                    # استثناء الأحجام الشائعة فقط إذا كانت متبوعة بـ ml/مل
                    pos = m.end()
                    after = text.lower()[pos:pos+5].strip()
                    if after.startswith('ml') or after.startswith('مل'):
                        continue  # هذا حجم
                    if v in {'212','360','1','2','3','4','5','6','7','8','9','11','12','13','14','15','16','17','18','19','21'}:
                        nums.add(v)
                return nums

            our_pnums = _extract_product_numbers(our_norm)
            c_pnums = _extract_product_numbers(self.norm_names[idx])
            if our_pnums and c_pnums and our_pnums != c_pnums:
                continue

            # ═══ مقارنة خط الإنتاج (الحل الجذري) ═══
            pline_penalty = 0
            if our_pline and c_pl:
                pl_score = fuzz.token_sort_ratio(our_pline, c_pl)
                if our_br and c_br:
                    # نفس الماركة → مقارنة خط الإنتاج صارمة جداً
                    # باروندا≠باردون(77%), الاباي≠اسبريت(75%)
                    # سوفاج=سوفاج(100%), عود مود=عود سيلك مود(85%)
                    if pl_score < 78:
                        continue  # رفض نهائي - خطوط إنتاج مختلفة
                    elif pl_score < 88:
                        pline_penalty = -20
                    elif pl_score < 94:
                        pline_penalty = -10
                else:
                    # ماركات مختلفة أو غير معروفة → مقارنة أكثر صرامة
                    if pl_score < 65:
                        pline_penalty = -35
                    elif pl_score < 80:
                        pline_penalty = -22

            # ═══ score تفصيلي ═══
            n1, n2 = our_norm, self.norm_names[idx]
            s1 = fuzz.token_sort_ratio(n1, n2)
            s2 = fuzz.token_set_ratio(n1, n2)
            s3 = fuzz.partial_ratio(n1, n2)
            base = s1*0.35 + s2*0.35 + s3*0.30

            # ═══ تعديلات الماركة ═══
            if our_br and c_br:
                base += 10 if normalize(our_br)==normalize(c_br) else -25
            elif our_br and not c_br:
                base -= 25  # منتجنا له ماركة لكن المنافس بدون → خصم كبير
            elif not our_br and c_br:
                base -= 25  # العكس
            elif not our_br and not c_br:
                # كلاهما بدون ماركة → خصم لأن المطابقة غير موثوقة
                base -= 10

            # ═══ تعديلات الحجم ═══
            if our_sz > 0 and c_sz > 0:
                d = abs(our_sz - c_sz)
                base += 10 if d==0 else (-5 if d<=5 else -18 if d<=20 else -30)
            if our_tp and c_tp and our_tp != c_tp: base -= 14
            if our_gd and c_gd and our_gd != c_gd:
                continue  # رفض نهائي - رجالي ≠ نسائي
            elif (our_gd or c_gd) and our_gd != c_gd:
                base -= 15  # أحدهما محدد والآخر فارغ

            # ═══ تطبيق عقوبة خط الإنتاج ═══
            base += pline_penalty

            score = round(max(0, min(100, base)), 1)
            if score < MATCH_THRESHOLD: continue

            seen.add(name)
            cands.append({
                "name": name, "score": score,
                "price": self.prices[idx], "product_id": self.ids[idx],
                "brand": c_br, "size": c_sz, "type": c_tp, "gender": c_gd,
                "competitor": self.comp_name,
            })

        cands.sort(key=lambda x: x["score"], reverse=True)
        return cands[:top_n]


# ═══════════════════════════════════════════════════════
#  Gemini Batch — 10 منتجات / استدعاء
# ═══════════════════════════════════════════════════════
_GURL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"

def _ai_batch(batch):
    """
    batch: [{"our":str, "price":float, "candidates":[...]}]
    → [int]  (0-based index | -1=no match)
    """
    if not GEMINI_API_KEYS or not batch: return [0]*len(batch)

    # cache key
    ck = hashlib.md5(json.dumps(
        [{"o":x["our"], "c":[c["name"] for c in x["candidates"]]} for x in batch],
        ensure_ascii=False, sort_keys=True).encode()).hexdigest()
    cached = _cget(ck)
    if cached is not None: return cached

    lines = []
    for i, it in enumerate(batch):
        cands = "\n".join(
            f"  {j+1}. {c['name']} | {int(c.get('size',0))}ml | "
            f"{c.get('type','?')} | {c.get('gender','?')} | {c.get('price',0):.0f}ر.س"
            for j,c in enumerate(it["candidates"])
        )
        lines.append(f"[{i+1}] منتجنا: «{it['our']}» ({it['price']:.0f}ر.س)\n{cands}")

    prompt = (
        "خبير عطور فاخرة. لكل منتج اختر رقم المرشح المطابق تماماً أو 0 إذا لا يوجد.\n"
        "الشروط: ✓نفس الماركة ✓نفس الحجم (±5ml) ✓نفس EDP/EDT ✓نفس الجنس إذا مذكور\n\n"
        + "\n\n".join(lines)
        + f'\n\nJSON فقط: {{"results":[r1,r2,...,r{len(batch)}]}}'
    )

    payload = {"contents":[{"parts":[{"text":prompt}]}],
               "generationConfig":{"temperature":0,"maxOutputTokens":200,"topP":1,"topK":1}}

    for attempt in range(3):
        for key in GEMINI_API_KEYS:
            if not key: continue
            try:
                r = _req.post(f"{_GURL}?key={key}", json=payload, timeout=22)
                if r.status_code == 200:
                    txt = r.json()["candidates"][0]["content"]["parts"][0]["text"]
                    clean = re.sub(r'```json|```','',txt).strip()
                    s = clean.find('{'); e = clean.rfind('}')+1
                    if s>=0 and e>s:
                        raw = json.loads(clean[s:e]).get("results",[])
                        out = []
                        for j,it in enumerate(batch):
                            n = raw[j] if j<len(raw) else 1
                            try: n=int(n)
                            except: n=1
                            if 1<=n<=len(it["candidates"]): out.append(n-1)
                            elif n==0: out.append(-1)
                            else: out.append(0)
                        _cset(ck, out)
                        return out
                elif r.status_code==429:
                    time.sleep(2**attempt)
            except: continue
        time.sleep(1)
    return [0]*len(batch)


# ═══════════════════════════════════════════════════════
#  بناء صف النتيجة
# ═══════════════════════════════════════════════════════
def _row(product, our_price, our_id, brand, size, ptype, gender,
         best=None, override=None, src="", all_cands=None):
    sz_str = f"{int(size)}ml" if size else ""
    if best is None:
        return dict(المنتج=product, معرف_المنتج=our_id, السعر=our_price,
                    الماركة=brand, الحجم=sz_str, النوع=ptype, الجنس=gender,
                    منتج_المنافس="—", معرف_المنافس="", سعر_المنافس=0,
                    الفرق=0, نسبة_التطابق=0, ثقة_AI="—",
                    القرار=override or "🔍 منتجات مفقودة",
                    الخطورة="", المنافس="", عدد_المنافسين=0,
                    جميع_المنافسين=[], مصدر_المطابقة=src or "—",
                    تاريخ_المطابقة=datetime.now().strftime("%Y-%m-%d"))

    cp    = float(best.get("price") or 0)
    score = float(best.get("score") or 0)
    diff  = round(our_price - cp, 2) if (our_price>0 and cp>0) else 0
    # نظام الخطورة حسب AI_COMPARISON_INSTRUCTIONS (نسبة مئوية + ثقة)
    diff_pct = abs((diff / cp) * 100) if cp > 0 else 0
    if diff_pct > 20 and score >= 85:
        risk = "🔴 حرج"
    elif diff_pct > 10 and score >= 75:
        risk = "🟡 متوسط"
    else:
        risk = "🟢 منخفض"

    # ═══ توزيع النتائج على الأقسام ═══
    # 🔴 سعر أعلى: سعرنا أعلى من المنافس بأكثر من 10 ريال
    # 🟢 سعر أقل: سعرنا أقل من المنافس بأكثر من 10 ريال
    # ✅ موافق: سعرنا مناسب (فرق ≤ 10 ريال)
    # ⚠️ تحت المراجعة: المطابقة غير مؤكدة (ثقة منخفضة)
    PRICE_DIFF_THRESHOLD = 10  # فرق السعر المقبول بالريال
    if override:
        dec = override
    elif src in ("gemini","auto") or score >= HIGH_CONFIDENCE:
        # مطابقة مؤكدة → توزيع حسب السعر
        if our_price > 0 and cp > 0:
            if diff > PRICE_DIFF_THRESHOLD:     dec = "🔴 سعر أعلى"
            elif diff < -PRICE_DIFF_THRESHOLD:   dec = "🟢 سعر أقل"
            else:                                dec = "✅ موافق"
        else:
            dec = "⚠️ تحت المراجعة"  # لا يوجد سعر → مراجعة
    elif score >= REVIEW_THRESHOLD:
        # مطابقة محتملة لكن تحتاج تأكيد → تحت المراجعة
        dec = "⚠️ تحت المراجعة"
    else:
        dec = "⚠️ تحت المراجعة"

    ai_lbl = {"gemini":f"🤖✅({score:.0f}%)",
              "auto":f"🎯({score:.0f}%)",
              "gemini_no_match":"🤖❌"}.get(src, f"{score:.0f}%")

    ac = (all_cands or [best])[:5]
    return dict(المنتج=product, معرف_المنتج=our_id, السعر=our_price,
                الماركة=brand, الحجم=sz_str, النوع=ptype, الجنس=gender,
                منتج_المنافس=best["name"], معرف_المنافس=best.get("product_id",""),
                سعر_المنافس=cp, الفرق=diff, نسبة_التطابق=score, ثقة_AI=ai_lbl,
                القرار=dec, الخطورة=risk, المنافس=best.get("competitor",""),
                عدد_المنافسين=len({c.get("competitor","") for c in ac}),
                جميع_المنافسين=ac, مصدر_المطابقة=src or "fuzzy",
                تاريخ_المطابقة=datetime.now().strftime("%Y-%m-%d"))


# ═══════════════════════════════════════════════════════
#  التحليل الكامل — v21 الهجين الفائق السرعة
# ═══════════════════════════════════════════════════════
def run_full_analysis(our_df, comp_dfs, progress_callback=None, use_ai=True):
    """
    1. بناء CompIndex لكل منافس (تطبيع مسبق)
    2. لكل منتجنا → search vectorized
    3. score≥97 → تلقائي | 62-96 → AI batch | <62 → مفقود
    """
    results = []
    our_col       = _fcol(our_df, ["المنتج","اسم المنتج","Product","Name","name"])
    our_price_col = _fcol(our_df, ["السعر","سعر","Price","price","PRICE"])
    our_id_col    = _fcol(our_df, [
        "رقم المنتج","معرف المنتج","المعرف","معرف","رقم_المنتج","معرف_المنتج",
        "product_id","Product ID","Product_ID","ID","id","Id",
        "SKU","sku","Sku","رمز المنتج","رمز_المنتج","رمز المنتج sku",
        "الكود","كود","Code","code","الرقم","رقم","Barcode","barcode","الباركود"
    ])

    # ── بناء الفهارس المسبقة ──
    indices = {}
    for cname, cdf in comp_dfs.items():
        ccol = _fcol(cdf, ["المنتج","اسم المنتج","Product","Name","name"])
        icol = _fcol(cdf, [
            "رقم المنتج","معرف المنتج","المعرف","معرف","رقم_المنتج","معرف_المنتج",
            "product_id","Product ID","Product_ID","ID","id","Id",
            "SKU","sku","Sku","رمز المنتج","رمز_المنتج","رمز المنتج sku",
            "الكود","كود","Code","code","الرقم","رقم","Barcode","barcode","الباركود"
        ])
        indices[cname] = CompIndex(cdf, ccol, icol, cname)

    total   = len(our_df)
    pending = []
    BATCH   = 12  # زيادة الـ batch لتقليل استدعاءات API

    def _flush():
        if not pending: return
        idxs = _ai_batch(pending)
        for j, it in enumerate(pending):
            ci = idxs[j] if j<len(idxs) else 0
            if ci < 0:
                results.append(_row(it["product"],it["our_price"],it["our_id"],
                                    it["brand"],it["size"],it["ptype"],it["gender"],
                                    None,"🔍 منتجات مفقودة","gemini_no_match"))
            else:
                best = it["candidates"][ci]
                results.append(_row(it["product"],it["our_price"],it["our_id"],
                                    it["brand"],it["size"],it["ptype"],it["gender"],
                                    best,src="gemini",all_cands=it["all_cands"]))
        pending.clear()

    for i, (_, row) in enumerate(our_df.iterrows()):
        product = str(row.get(our_col,"")).strip()
        if not product or is_sample(product):
            if progress_callback: progress_callback((i+1)/total)
            continue

        our_price = 0.0
        if our_price_col:
            try: our_price = float(str(row[our_price_col]).replace(",",""))
            except: pass

        our_id  = _pid(row, our_id_col)
        brand   = extract_brand(product)
        size    = extract_size(product)
        ptype   = extract_type(product)
        gender  = extract_gender(product)
        our_n   = normalize(product)
        our_pl  = extract_product_line(product, brand)

        # ── جمع المرشحين من كل الفهارس ──
        all_cands = []
        for idx_obj in indices.values():
            all_cands.extend(idx_obj.search(our_n, brand, size, ptype, gender, our_pline=our_pl, top_n=5))

        if not all_cands:
            results.append(_row(product,our_price,our_id,brand,size,ptype,gender,
                                None,"🔍 منتجات مفقودة"))
            if progress_callback: progress_callback((i+1)/total)
            continue

        all_cands.sort(key=lambda x: x["score"], reverse=True)
        top5  = all_cands[:5]
        best0 = top5[0]

        if best0["score"] >= 97 or not use_ai:
            # واضح تماماً → لا حاجة AI
            results.append(_row(product,our_price,our_id,brand,size,ptype,gender,
                                best0,src="auto",all_cands=all_cands))
        else:
            # غامض → AI batch
            pending.append(dict(product=product,our_price=our_price,our_id=our_id,
                                brand=brand,size=size,ptype=ptype,gender=gender,
                                candidates=top5,all_cands=all_cands,
                                our=product,price=our_price))
            if len(pending) >= BATCH: _flush()

        if progress_callback: progress_callback((i+1)/total)

    _flush()
    return pd.DataFrame(results)


# ═══════════════════════════════════════════════════════
#  المنتجات المفقودة (محدثة - الإصدار المتوازن والدقيق)
# ═══════════════════════════════════════════════════════
def find_missing_products(our_df, comp_dfs):
    our_col  = _fcol(our_df, ["المنتج","اسم المنتج","Product","Name","name"])

    # تجهيز بيانات منتجاتنا للبحث السريع
    our_items = []
    for _, r in our_df.iterrows():
        name = str(r.get(our_col, "")).strip()
        if not name or is_sample(name): continue
        brand = extract_brand(name)
        our_items.append({
            "norm": normalize(name),
            "brand": brand,
            "norm_brand": normalize(brand) if brand else "",
            "pline": extract_product_line(name, brand),
            "size": extract_size(name),
            "type": extract_type(name),
            "gender": extract_gender(name)
        })

    missing, seen = [], set()
    for cname, cdf in comp_dfs.items():
        ccol = _fcol(cdf, ["المنتج","اسم المنتج","Product","Name","name"])
        icol = _fcol(cdf, [
            "رقم المنتج","معرف المنتج","المعرف","معرف","رقم_المنتج","معرف_المنتج",
            "product_id","Product ID","Product_ID","ID","id","Id",
            "SKU","sku","Sku","رمز المنتج","رمز_المنتج","رمز المنتج sku",
            "الكود","كود","Code","code","الرقم","رقم","Barcode","barcode","الباركود"
        ])

        for _, row in cdf.iterrows():
            cp = str(row.get(ccol, "")).strip()
            if not cp or is_sample(cp): continue
            cn = normalize(cp)
            if not cn: continue

            c_brand = extract_brand(cp)
            c_pline = extract_product_line(cp, c_brand)
            c_size = extract_size(cp)
            c_type = extract_type(cp)
            c_gender = extract_gender(cp)

            # فلتر ماركة مرن: إذا عرفنا ماركة المنافس، نبحث في منتجاتنا
            # التي تحمل نفس الماركة + المنتجات التي لم يُتعرف على ماركتها
            if c_brand:
                c_brand_norm = normalize(c_brand)
                candidates = [o for o in our_items
                              if o["norm_brand"] == c_brand_norm or o["brand"] == ""]
            else:
                candidates = our_items

            is_missing = True
            if candidates:
                norms = [c["norm"] for c in candidates]

                # البحث باستخدام token_sort_ratio و token_set_ratio معاً
                matches_sort = rf_process.extract(cn, norms, scorer=fuzz.token_sort_ratio, limit=5)
                matches_set  = rf_process.extract(cn, norms, scorer=fuzz.token_set_ratio,  limit=5)

                # دمج النتائج: لكل مرشح نأخذ أعلى قيمة من الخوارزميتين
                best_by_idx = {}
                for match_norm, match_score, match_idx in matches_sort:
                    if match_idx not in best_by_idx or match_score > best_by_idx[match_idx]:
                        best_by_idx[match_idx] = match_score
                for match_norm, match_score, match_idx in matches_set:
                    if match_idx not in best_by_idx or match_score > best_by_idx[match_idx]:
                        best_by_idx[match_idx] = match_score

                # ترتيب المرشحين حسب أعلى سكور
                sorted_candidates = sorted(best_by_idx.items(), key=lambda x: x[1], reverse=True)

                for match_idx, match_score in sorted_candidates:
                    if match_score < 60:
                        break  # لا فائدة من فحص مرشحين بسكور منخفض جداً
                    matched_item = candidates[match_idx]
                    penalty = 0

                    # تطبيق عقوبات في حال اختلاف المواصفات الجوهرية
                    if c_size > 0 and matched_item["size"] > 0 and abs(c_size - matched_item["size"]) > 10:
                        penalty += 25  # حجم مختلف
                    if c_type and matched_item["type"] and c_type != matched_item["type"]:
                        penalty += 15  # تركيز مختلف (EDP vs EDT)
                    if c_gender and matched_item["gender"] and c_gender != matched_item["gender"]:
                        penalty += 25  # جنس مختلف

                    if c_pline and matched_item["pline"]:
                        pl_score = fuzz.token_sort_ratio(c_pline, matched_item["pline"])
                        if pl_score < 75:
                            penalty += 20  # خط إنتاج مختلف

                    final_score = match_score - penalty

                    # إذا بقي السكور أعلى من 80 بعد العقوبات، إذن المنتج موجود لدينا فعلاً
                    if final_score > 80:
                        is_missing = False
                        break

            if is_missing:
                seen.add(cn)
                missing.append({
                    "منتج_المنافس": cp, "معرف_المنافس": _pid(row, icol),
                    "سعر_المنافس": _price(row), "المنافس": cname,
                    "الماركة": c_brand,
                    "الحجم": f"{int(c_size)}ml" if c_size else "",
                    "النوع": c_type, "الجنس": c_gender,
                    "تاريخ_الرصد": datetime.now().strftime("%Y-%m-%d"),
                })

    return pd.DataFrame(missing) if missing else pd.DataFrame()
# ═══════════════════════════════════════════════════════
#  تصدير Excel ملوّن
# ═══════════════════════════════════════════════════════
def export_excel(df, sheet_name="النتائج"):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    edf = df.copy()
    for col in ["جميع المنافسين","جميع_المنافسين"]:
        if col in edf.columns: edf = edf.drop(columns=[col])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        edf.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        ws = writer.sheets[sheet_name[:31]]
        hfill = PatternFill("solid", fgColor="1a1a2e")
        hfont = Font(color="FFFFFF", bold=True, size=10)
        for cell in ws[1]:
            cell.fill=hfill; cell.font=hfont
            cell.alignment=Alignment(horizontal="center")
        # تم تعديل المسميات هنا لمطابقة طلبك بدقة تامة
        COLORS = {"🔴 سعر أعلى":"FFCCCC","🟢 سعر أقل":"CCFFCC",
                  "✅ موافق":"CCFFEE","⚠️ تحت المراجعة":"FFF3CC","🔍 منتجات مفقودة":"CCE5FF"}
        dcol = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value and "القرار" in str(cell.value): dcol=i; break
        if dcol:
            for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
                val = str(ws.cell(ri,dcol).value or "")
                for k,c in COLORS.items():
                    if k.split()[0] in val:
                        for cell in row: cell.fill=PatternFill("solid",fgColor=c)
                        break
        for ci, col in enumerate(ws.columns, 1):
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(ci)].width = min(w+4, 55)
    return output.getvalue()

def export_section_excel(df, sname):
    return export_excel(df, sheet_name=sname[:31])
